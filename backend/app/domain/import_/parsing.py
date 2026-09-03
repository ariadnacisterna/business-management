import csv
import io
from dataclasses import dataclass

from openpyxl import load_workbook

from app.constants.import_ import (
    ALL_COLUMNS,
    CSV_EXTENSION,
    EXCEL_EXTENSION,
    FIRST_DATA_ROW_NUMBER,
    MAX_DATA_ROWS,
    MAX_FILE_SIZE_BYTES,
    REQUIRED_COLUMNS,
)
from app.domain.import_.errors import (
    EmptyFile,
    FileTooLarge,
    MissingColumns,
    TooManyRows,
    UnsupportedFileType,
)


@dataclass
class ParsedRow:
    row_number: int
    values: dict[str, str]


def _parse_csv(content: bytes) -> tuple[list[dict[str, str | None]], list[str]]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    return list(reader), list(reader.fieldnames or [])


def _cell_to_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _parse_excel(content: bytes) -> tuple[list[dict[str, str | None]], list[str]]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    rows_iter = sheet.iter_rows(values_only=True)

    try:
        header_row = next(rows_iter)
    except StopIteration:
        return [], []
    header = [_cell_to_text(cell).strip() for cell in header_row]

    rows: list[dict[str, str | None]] = []
    for raw_row in rows_iter:
        if all(cell is None for cell in raw_row):
            continue
        rows.append(
            {column: _cell_to_text(cell) for column, cell in zip(header, raw_row, strict=False)}
        )
    return rows, header


def parse_file(filename: str, content: bytes) -> list[ParsedRow]:
    if len(content) > MAX_FILE_SIZE_BYTES:
        raise FileTooLarge(f"El archivo supera el tamano maximo de {MAX_FILE_SIZE_BYTES} bytes")

    lowered = filename.lower()
    if lowered.endswith(CSV_EXTENSION):
        raw_rows, header = _parse_csv(content)
    elif lowered.endswith(EXCEL_EXTENSION):
        raw_rows, header = _parse_excel(content)
    else:
        raise UnsupportedFileType("Solo se admiten archivos CSV (.csv) o Excel (.xlsx)")

    header_names = {(name or "").strip() for name in header}
    missing = [column for column in REQUIRED_COLUMNS if column not in header_names]
    if missing:
        raise MissingColumns(missing)

    if not raw_rows:
        raise EmptyFile("El archivo no tiene filas de datos")

    if len(raw_rows) > MAX_DATA_ROWS:
        raise TooManyRows(f"El archivo supera el maximo de {MAX_DATA_ROWS} filas")

    parsed_rows = []
    for index, raw in enumerate(raw_rows):
        row_number = FIRST_DATA_ROW_NUMBER + index
        values = {column: (raw.get(column) or "").strip() for column in ALL_COLUMNS}
        parsed_rows.append(ParsedRow(row_number=row_number, values=values))
    return parsed_rows
